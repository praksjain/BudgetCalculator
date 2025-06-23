import os
import logging
import re
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import PyPDF2
import docx
from sqlalchemy.orm import Session
from io import BytesIO
from ..core.config import settings
from ..models.rfp import RFP
from ..models.rfp_analysis import RFPAnalysis, AnalysisTask, AnalysisSubtask

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Optional imports for AI services
try:
    import openai
    OPENAI_AVAILABLE = True
    logger.info("✅ OpenAI package available")
except ImportError:
    OPENAI_AVAILABLE = False
    logger.warning("❌ OpenAI package not available")

try:
    import requests
    REQUESTS_AVAILABLE = True
    logger.info("✅ Requests package available")
except ImportError:
    REQUESTS_AVAILABLE = False
    logger.warning("❌ Requests package not available")

try:
    import google.generativeai as genai
    GOOGLE_AI_AVAILABLE = True
    logger.info("✅ Google Generative AI package available")
except ImportError:
    GOOGLE_AI_AVAILABLE = False
    logger.warning("❌ Google Generative AI package not available")

class DocumentAnalysisService:
    def __init__(self):
        self.hourly_rate = 100.0  # Default hourly rate
        if settings.OPENAI_API_KEY:
            openai.api_key = settings.OPENAI_API_KEY
    
    def extract_text_from_file(self, file_path: str, file_type: str) -> str:
        """Extract text from various document formats"""
        try:
            if file_type.lower() == 'pdf':
                return self._extract_from_pdf(file_path)
            elif file_type.lower() in ['doc', 'docx']:
                return self._extract_from_docx(file_path)
            else:
                # For other file types, try to read as plain text
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            raise
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF using PyPDF2"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            logger.error(f"Error reading PDF {file_path}: {str(e)}")
            raise
        return text
    
    def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX using python-docx"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error reading DOCX {file_path}: {str(e)}")
            raise

    async def analyze_rfp_document(self, db: Session, rfp_id: int, model: str = "gpt-3.5-turbo", application_type: str = "web", technology_config: dict = None, rate_card = None) -> RFPAnalysis:
        """
        Analyze RFP document and generate comprehensive RFP Summary
        
        Args:
            db: Database session
            rfp_id: ID of the RFP to analyze
            model: AI model to use for analysis
            application_type: Type of application (web, mobile, both)
            technology_config: Technology configuration for different components
            rate_card: Rate card for different roles
            
        Returns:
            RFPAnalysis: The analysis results stored in database
        """
        
        # Convert technology_config to dict if it's a Pydantic model, or set default if None
        if technology_config is None:
            technology_config_dict = {
                "frontend": "react",
                "backend": "python",
                "database": "postgresql",
                "cloud": "aws"
            }
        elif hasattr(technology_config, 'dict'):
            technology_config_dict = technology_config.dict()
        else:
            technology_config_dict = technology_config
        
        # Convert rate_card to dict if it's a Pydantic model, or set default if None
        if rate_card is None:
            rate_card_dict = {
                "senior_developer": 90.0,
                "mid_developer": 75.0,
                "junior_developer": 55.0,
                "tech_lead": 110.0,
                "project_manager": 85.0,
                "business_analyst": 70.0,
                "ui_ux_designer": 80.0,
                "qa_engineer": 65.0,
                "devops_engineer": 95.0,
                "blockchain_developer": 120.0,
                "smart_contract_auditor": 150.0
            }
        elif hasattr(rate_card, 'dict'):
            rate_card_dict = rate_card.dict()
        else:
            rate_card_dict = rate_card
        
        print(f"\n🔍 STARTING RFP ANALYSIS for RFP ID: {rfp_id}")
        print(f"📱 Application Type: {application_type}")
        print(f"🛠️ Technology Config: {technology_config}")
        print(f"💰 Rate Card: Average rate ${sum(rate_card_dict.values()) / len(rate_card_dict):.0f}/hr")
        logger.info(f"Starting RFP analysis for RFP ID: {rfp_id}")
        
        # Get RFP from database
        rfp = db.query(RFP).filter(RFP.id == rfp_id).first()
        if not rfp:
            raise ValueError("RFP not found")
        
        if not rfp.document_path or not os.path.exists(rfp.document_path):
            raise ValueError("RFP document not found")
        
        print(f"📄 Found RFP: {rfp.title}")
        print(f"📁 Document path: {rfp.document_path}")
        
        # Determine file type from extension
        file_extension = os.path.splitext(rfp.document_path)[1].lower()
        file_type = file_extension.replace('.', '')
        
        # Extract text from document
        print(f"📝 Extracting text from {file_type.upper()} document...")
        document_text = self.extract_text_from_file(rfp.document_path, file_type)
        
        print(f"✅ Extracted {len(document_text)} characters from document")
        logger.info(f"Extracted {len(document_text)} characters from document")
        
        # Generate RFP Summary using AI
        print(f"🤖 Generating RFP Summary using {model}...")
        rfp_summary = await self._generate_rfp_summary(document_text, model, application_type, technology_config)
        
        # Parse the AI response to extract different sections
        print(f"🔍 Parsing AI response to extract analysis sections...")
        parsed_analysis = self._parse_analysis_sections(rfp_summary)
        
        # Calculate estimated hours and cost based on content complexity
        estimated_hours = float(self._estimate_project_hours(document_text, application_type))
        estimated_cost = float(self._calculate_project_cost(estimated_hours, rate_card_dict))
        
        # Check if analysis already exists
        existing_analysis = db.query(RFPAnalysis).filter(RFPAnalysis.rfp_id == rfp_id).first()
        if existing_analysis:
            print(f"🔄 Updating existing analysis...")
            # Update existing analysis with parsed data
            tech_stack_summary = f"{application_type.title()} app: {technology_config_dict['frontend']} + {technology_config_dict['backend']} + {technology_config_dict['database']} on {technology_config_dict['cloud']}"
            existing_analysis.summary = parsed_analysis.get('summary', rfp_summary)
            existing_analysis.scope = parsed_analysis.get('scope', 'Project scope analysis completed')
            existing_analysis.requirements = parsed_analysis.get('requirements', 'Key requirements extracted from RFP document')
            existing_analysis.deliverables = parsed_analysis.get('deliverables', 'Project deliverables identified from analysis')
            existing_analysis.timeline = parsed_analysis.get('timeline', f"{int(estimated_hours) // 40} to {(int(estimated_hours) // 40) + 4} weeks")
            existing_analysis.complexity_level = parsed_analysis.get('complexity', self._assess_complexity(document_text, estimated_hours))
            existing_analysis.technology_stack = tech_stack_summary
            existing_analysis.risks = parsed_analysis.get('risks', 'Risk assessment completed based on project requirements')
            existing_analysis.total_estimated_hours = estimated_hours
            existing_analysis.total_estimated_cost = estimated_cost
            existing_analysis.confidence_level = 0.8
            
            db.commit()
            db.refresh(existing_analysis)
            analysis = existing_analysis
        else:
            print(f"✨ Creating new analysis record...")
            # Create new analysis record with parsed data
            tech_stack_summary = f"{application_type.title()} app: {technology_config_dict['frontend']} + {technology_config_dict['backend']} + {technology_config_dict['database']} on {technology_config_dict['cloud']}"
            analysis = RFPAnalysis(
                rfp_id=rfp_id,
                summary=parsed_analysis.get('summary', rfp_summary),
                scope=parsed_analysis.get('scope', 'Project scope analysis completed'),
                requirements=parsed_analysis.get('requirements', 'Key requirements extracted from RFP document'),
                deliverables=parsed_analysis.get('deliverables', 'Project deliverables identified from analysis'),
                timeline=parsed_analysis.get('timeline', f"{int(estimated_hours) // 40} to {(int(estimated_hours) // 40) + 4} weeks"),
                complexity_level=parsed_analysis.get('complexity', self._assess_complexity(document_text, estimated_hours)),
                technology_stack=tech_stack_summary,
                risks=parsed_analysis.get('risks', 'Risk assessment completed based on project requirements'),
                total_estimated_hours=estimated_hours,
                total_estimated_cost=estimated_cost,
                confidence_level=0.8
            )
            
            db.add(analysis)
            db.commit()
            db.refresh(analysis)
        
        print(f"✅ RFP Analysis completed successfully!")
        print(f"📊 Analysis ID: {analysis.id}")
        logger.info(f"RFP analysis completed for RFP ID: {rfp_id}, Analysis ID: {analysis.id}")
        
        return analysis

    async def _generate_rfp_summary(self, document_text: str, model: str = "gpt-3.5-turbo", application_type: str = "web", technology_config = None) -> str:
        """
        Generate RFP Summary with 6 specific sections using AI
        
        Args:
            document_text: The extracted document content
            model: AI model to use for analysis
            application_type: Type of application (web, mobile, both)
            technology_config: Technology configuration for different components
            
        Returns:
            str: Formatted RFP summary with 6 sections
        """
        
        print(f"🔍 Analyzing document content ({len(document_text)} chars)...")
        
        # Try different AI services in order of preference
        ai_services = [
            ("Google Gemini", self._analyze_with_gemini),
            ("Hugging Face", self._analyze_with_huggingface), 
            ("OpenAI", self._analyze_with_openai)
        ]
        
        for service_name, service_method in ai_services:
            try:
                print(f"🤖 Trying {service_name}...")
                summary = await service_method(document_text, model, application_type, technology_config)
                print(f"✅ Successfully generated summary using {service_name}")
                return summary
            except Exception as e:
                print(f"❌ {service_name} failed: {str(e)}")
                logger.warning(f"{service_name} analysis failed: {str(e)}")
                continue
        
        # If all AI services fail, use enhanced mock analysis
        print(f"🔄 All AI services failed, using enhanced mock analysis...")
        logger.info("All AI services failed, falling back to mock analysis")
        return self._create_mock_rfp_summary(document_text)

    async def _analyze_with_gemini(self, document_text: str, model: str = "gpt-3.5-turbo", application_type: str = "web", technology_config = None) -> str:
        """Generate RFP Summary using Google Gemini"""
        
        if not GOOGLE_AI_AVAILABLE or not settings.GOOGLE_API_KEY:
            raise Exception("Google Gemini not available")
            
        print(f"🔑 Configuring Gemini API...")
        genai.configure(api_key=settings.GOOGLE_API_KEY)
        
        # Try multiple model names
        model_names = ['gemini-1.5-flash', 'gemini-1.5-pro', 'gemini-pro']
        
        for model_name in model_names:
            try:
                print(f"🔍 Trying model: {model_name}")
                model = genai.GenerativeModel(model_name)
                
                prompt = self._create_rfp_summary_prompt(document_text, application_type, technology_config)
                response = model.generate_content(prompt)
                
                if response.text:
                    print(f"✅ Received response from {model_name}")
                    return self._format_ai_response(response.text)
                    
            except Exception as e:
                print(f"❌ Model {model_name} failed: {str(e)}")
                continue
        
        raise Exception("No Gemini models available")

    async def _analyze_with_huggingface(self, document_text: str, model: str = "gpt-3.5-turbo", application_type: str = "web", technology_config = None) -> str:
        """Generate RFP Summary using Hugging Face"""
        
        if not REQUESTS_AVAILABLE or not settings.HUGGINGFACE_API_KEY:
            raise Exception("Hugging Face not available")
            
        # Use a text generation model for analysis
        API_URL = "https://api-inference.huggingface.co/models/microsoft/DialoGPT-large"
        headers = {"Authorization": f"Bearer {settings.HUGGINGFACE_API_KEY}"}
        
        prompt = self._create_rfp_summary_prompt(document_text, application_type, technology_config)
        payload = {
            "inputs": prompt,
            "parameters": {
                "max_length": 2000,
                "temperature": 0.3,
                "do_sample": True
            }
        }
        
        response = requests.post(API_URL, headers=headers, json=payload, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            if isinstance(result, list) and len(result) > 0:
                generated_text = result[0].get("generated_text", "")
                return self._format_ai_response(generated_text)
        
        raise Exception(f"Hugging Face API error: {response.status_code}")

    async def _analyze_with_openai(self, document_text: str, model: str = "gpt-3.5-turbo", application_type: str = "web", technology_config = None) -> str:
        """Generate RFP Summary using OpenAI"""
        
        if not OPENAI_AVAILABLE or not settings.OPENAI_API_KEY:
            raise Exception("OpenAI not available")
            
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
        
        prompt = self._create_rfp_summary_prompt(document_text, application_type, technology_config)
        
        response = await client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "You are an expert RFP analyst who creates comprehensive project summaries."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=2000,
            temperature=0.3
        )
        
        analysis_text = response.choices[0].message.content
        return self._format_ai_response(analysis_text)

    def _create_rfp_summary_prompt(self, document_text: str, application_type: str = "web", technology_config = None) -> str:
        """Create the specific prompt for RFP Summary generation"""
        
        # Convert technology_config to dict if it's a Pydantic model, or set default if None
        if technology_config is None:
            tech_config_dict = {
                "frontend": "react",
                "backend": "python",
                "database": "postgresql",
                "cloud": "aws"
            }
        elif hasattr(technology_config, 'dict'):
            tech_config_dict = technology_config.dict()
        else:
            tech_config_dict = technology_config
        
        # Create comprehensive tech description
        tech_description = f"{application_type.title()} application using {tech_config_dict['frontend']} frontend, {tech_config_dict['backend']} backend, {tech_config_dict['database']} database, deployed on {tech_config_dict['cloud']} cloud"
        
        # Add blockchain-specific context if blockchain technologies are detected
        blockchain_backends = ['ethereum', 'solana', 'polygon', 'hyperledger', 'cardano', 'polkadot']
        blockchain_storage = ['ipfs', 'arweave', 'filecoin', 'blockchain-native']
        blockchain_cloud = ['alchemy', 'infura', 'quicknode', 'chainlink', 'pinata']
        
        is_blockchain_project = (
            tech_config_dict['backend'] in blockchain_backends or
            tech_config_dict['database'] in blockchain_storage or
            tech_config_dict['cloud'] in blockchain_cloud
        )
        
        if is_blockchain_project:
            tech_description += "\n\n🔗 BLOCKCHAIN PROJECT DETECTED: This is a blockchain/Web3 project requiring specialized considerations for smart contracts, decentralization, gas optimization, security audits, and regulatory compliance."
        
        return f"""
        You are an expert RFP analyst. Please analyze the following RFP document and create a comprehensive RFP Summary.
        
        PROJECT CONFIGURATION:
        - Application Type: {application_type.title()}
        - Frontend Technology: {tech_config_dict['frontend']}
        - Backend Technology: {tech_config_dict['backend']}
        - Database: {tech_config_dict['database']}
        - Cloud Platform: {tech_config_dict['cloud']}
        
        PREFERRED TECHNOLOGY STACK: {tech_description}
        Consider this technology preference when analyzing technical requirements and constraints.

        RFP DOCUMENT CONTENT:
        {document_text}

        Please provide your analysis in EXACTLY the following format with these 6 sections:

        SUMMARY: [You MUST write EXACTLY 6 separate paragraphs with clear section headers. Follow this EXACT format:

        **EXECUTIVE OVERVIEW:**
        [Write 3-4 sentences describing the project's primary purpose, the requesting organization, and the strategic business objectives. Explain what problem this RFP solves and the expected organizational impact.]

        **FUNCTIONAL REQUIREMENTS:**
        [Write 3-4 sentences detailing the specific functional capabilities, features, user interactions, and core functionalities required. Include workflow requirements and user interface expectations.]

        **TECHNICAL REQUIREMENTS:**
        [Write 3-4 sentences outlining technical architecture, integration requirements, platform specifications, performance criteria, security requirements, and compliance standards. Include scalability and reliability expectations.]

        **OPERATIONAL REQUIREMENTS:**
        [Write 3-4 sentences describing the operational context including target user base, expected usage patterns, deployment requirements, maintenance expectations, and support requirements.]

        **BUSINESS CONTEXT & CONSTRAINTS:**
        [Write 3-4 sentences explaining business drivers, timeline constraints, budget considerations, regulatory requirements, and organizational policies that influence the project.]

        **SUCCESS CRITERIA & DELIVERABLES:**
        [Write 3-4 sentences defining measurable outcomes, acceptance criteria, key performance indicators, delivery milestones, and final deliverables that determine project success.]

        CRITICAL REQUIREMENTS:
        - You MUST include ALL 6 section headers with ** formatting
        - Each section must be 3-4 complete, well-structured sentences
        - Do not combine sections or skip any headers
        - Base your analysis on the actual document content provided
        - Write in professional, executive-level language
        - Focus on actionable insights and clear requirements

        Generate the summary now:
        """

    def _format_ai_response(self, ai_response: str) -> str:
        """Format and validate the AI response"""
        
        print(f"📝 Formatting AI response ({len(ai_response)} chars)...")
        
        # Check if response contains the required sections
        required_sections = [
            "**EXECUTIVE OVERVIEW:**",
            "**FUNCTIONAL REQUIREMENTS:**", 
            "**TECHNICAL REQUIREMENTS:**",
            "**OPERATIONAL REQUIREMENTS:**",
            "**BUSINESS CONTEXT & CONSTRAINTS:**",
            "**SUCCESS CRITERIA & DELIVERABLES:**"
        ]
        
        missing_sections = []
        for section in required_sections:
            if section not in ai_response:
                missing_sections.append(section)
        
        if missing_sections:
            print(f"⚠️ AI response missing sections: {missing_sections}")
            logger.warning(f"AI response missing sections: {missing_sections}")
            # Force enhancement to add missing sections
            return self._enhance_incomplete_response(ai_response)
        
        print(f"✅ AI response contains all required sections")
        return ai_response.strip()

    def _enhance_incomplete_response(self, partial_response: str) -> str:
        """Enhance incomplete AI response with missing sections"""
        
        print(f"🔧 Enhancing incomplete AI response...")
        
        # Start with what we have
        enhanced_response = partial_response.strip()
        
        # Add missing sections with placeholder content
        required_sections = {
            "**EXECUTIVE OVERVIEW:**": "This RFP outlines a comprehensive project initiative designed to address specific organizational needs and strategic objectives. The requesting organization seeks to implement a solution that will significantly improve operational efficiency and service delivery. The project represents a critical investment in technology infrastructure and process optimization. Success will drive measurable improvements in organizational performance and stakeholder satisfaction.",
            
            "**FUNCTIONAL REQUIREMENTS:**": "The solution must provide comprehensive functionality supporting core business processes and user workflows. Key capabilities include user management, data processing, reporting, and integration with existing systems. The system should offer intuitive user interfaces with role-based access controls and automated workflow management. Real-time processing and responsive design across multiple platforms are essential requirements.",
            
            "**TECHNICAL REQUIREMENTS:**": "The technical architecture must support scalable, secure, and reliable operations with modern technology standards. Integration capabilities with existing enterprise systems and third-party services are required. Performance specifications include high availability, data security, and compliance with industry standards. The solution should leverage cloud-native technologies with appropriate backup and disaster recovery capabilities.",
            
            "**OPERATIONAL REQUIREMENTS:**": "The system will serve multiple user groups with varying access levels and operational responsibilities. Deployment must accommodate existing infrastructure while minimizing disruption to current operations. Ongoing maintenance and support requirements include monitoring, updates, and user assistance. Training and change management support will be necessary for successful implementation and adoption.",
            
            "**BUSINESS CONTEXT & CONSTRAINTS:**": "This initiative operates within specific budget parameters and timeline constraints established by organizational priorities. Regulatory compliance and governance requirements must be maintained throughout implementation and operations. The project aligns with broader digital transformation goals and strategic business objectives. Resource allocation and stakeholder coordination will be critical success factors.",
            
            "**SUCCESS CRITERIA & DELIVERABLES:**": "Project success will be measured through specific performance metrics, user adoption rates, and operational efficiency improvements. Key deliverables include fully functional system, comprehensive documentation, user training, and ongoing support arrangements. Acceptance criteria focus on meeting functional requirements, performance benchmarks, and user satisfaction targets. Final delivery includes production deployment with validated performance and compliance standards."
        }
        
        # Check and add missing sections
        for section_header, default_content in required_sections.items():
            if section_header not in enhanced_response:
                enhanced_response += f"\n\n{section_header}\n{default_content}"
                print(f"   ➕ Added missing section: {section_header}")
        
        return enhanced_response

    def _create_mock_rfp_summary(self, document_text: str) -> str:
        """Create a mock RFP summary when AI services are unavailable"""
        
        print(f"🛠️ Creating enhanced mock RFP summary...")
        logger.info("Creating mock RFP summary based on document analysis")
        
        # Analyze document content for context
        text_lower = document_text.lower()
        word_count = len(document_text.split())
        
        # Detect project characteristics
        project_type = self._detect_project_type(text_lower)
        features = self._detect_features(text_lower)
        # Convert word_count to int to ensure compatibility
        complexity = self._assess_complexity_by_features(int(word_count), features)
        
        print(f"📊 Document analysis:")
        print(f"   - Project type: {project_type}")
        print(f"   - Word count: {word_count}")
        print(f"   - Features detected: {len(features)}")
        print(f"   - Complexity: {complexity}")
        
        return f"""**EXECUTIVE OVERVIEW:**
This RFP outlines a comprehensive {project_type.lower()} initiative designed to address critical organizational needs and strategic business objectives. The requesting organization seeks to implement a modern solution that will significantly enhance operational efficiency, improve service delivery, and support digital transformation goals. The project represents a strategic investment in technology infrastructure that will enable scalable growth and competitive advantage. Success will drive measurable improvements in organizational performance, user satisfaction, and operational metrics.

**FUNCTIONAL REQUIREMENTS:**
The solution must provide comprehensive functionality supporting core business processes including {', '.join(features[:3]) if features else 'user management, data processing, and reporting'}. Key capabilities include intuitive user interfaces, automated workflow management, real-time data processing, and comprehensive reporting and analytics. The system should offer role-based access controls, multi-user collaboration features, and seamless integration with existing business processes. Mobile accessibility and responsive design across multiple platforms are essential for supporting diverse user needs and usage patterns.

**TECHNICAL REQUIREMENTS:**
The technical architecture must support {complexity} complexity operations with modern, scalable technology standards and cloud-native infrastructure. Integration capabilities with existing enterprise systems, third-party services, and external APIs are required to ensure seamless data flow and operational continuity. Performance specifications include high availability (99.9% uptime), robust security protocols, data encryption, and compliance with industry standards and regulations. The solution should leverage modern development frameworks with appropriate backup, disaster recovery, and monitoring capabilities.

**OPERATIONAL REQUIREMENTS:**
The system will serve multiple user groups including end-users, administrators, and stakeholders with varying access levels and operational responsibilities. Deployment must accommodate existing organizational infrastructure while minimizing disruption to current operations and maintaining business continuity. Daily operations require intuitive interfaces, efficient workflows, and responsive support systems to ensure optimal user adoption and satisfaction. Ongoing maintenance includes regular updates, security patches, performance monitoring, and comprehensive user support and training programs.

**BUSINESS CONTEXT & CONSTRAINTS:**
This initiative operates within specific organizational budget parameters and timeline constraints established by strategic business priorities and competitive market pressures. Regulatory compliance requirements, data privacy standards, and governance policies must be maintained throughout implementation and ongoing operations. The project aligns with broader digital transformation initiatives and long-term strategic business objectives focused on growth and operational excellence. Resource allocation, stakeholder coordination, and change management will be critical success factors requiring executive sponsorship and cross-functional collaboration.

**SUCCESS CRITERIA & DELIVERABLES:**
Project success will be measured through specific performance metrics including user adoption rates (target >80%), operational efficiency improvements (target >30%), and system performance benchmarks (response time <2 seconds). Key deliverables include fully functional production system, comprehensive technical documentation, user training materials, and ongoing support arrangements with defined service level agreements. Acceptance criteria focus on meeting all functional requirements, achieving performance targets, passing security audits, and maintaining user satisfaction scores above 85%. Final delivery includes production deployment with validated performance, comprehensive testing results, and documented compliance with all specified requirements and standards."""

    def _detect_project_type(self, text_lower: str) -> str:
        """Detect the type of project from document content"""
        
        project_indicators = {
            "Web Application": ["web", "website", "portal", "dashboard", "online"],
            "Mobile Application": ["mobile", "app", "ios", "android", "smartphone"],
            "Enterprise System": ["enterprise", "erp", "crm", "system integration"],
            "E-commerce Platform": ["ecommerce", "e-commerce", "online store", "shopping"],
            "Data Platform": ["data", "analytics", "reporting", "business intelligence"],
            "Custom Software": ["software", "application", "system", "platform"]
        }
        
        for project_type, keywords in project_indicators.items():
            if any(keyword in text_lower for keyword in keywords):
                return project_type
        
        return "Custom Software Solution"
    
    def _parse_analysis_sections(self, ai_response: str) -> dict:
        """Parse the AI response to extract different analysis sections"""
        
        sections = {}
        
        # Define section patterns to extract
        section_patterns = {
            'summary': r'\*\*EXECUTIVE OVERVIEW:\*\*(.*?)(?=\*\*[A-Z]|\Z)',
            'requirements': r'\*\*FUNCTIONAL REQUIREMENTS:\*\*(.*?)(?=\*\*[A-Z]|\Z)',
            'scope': r'\*\*TECHNICAL REQUIREMENTS:\*\*(.*?)(?=\*\*[A-Z]|\Z)',
            'deliverables': r'\*\*SUCCESS CRITERIA & DELIVERABLES:\*\*(.*?)(?=\*\*[A-Z]|\Z)',
            'risks': r'\*\*BUSINESS CONTEXT & CONSTRAINTS:\*\*(.*?)(?=\*\*[A-Z]|\Z)',
            'timeline': r'\*\*OPERATIONAL REQUIREMENTS:\*\*(.*?)(?=\*\*[A-Z]|\Z)'
        }
        
        import re
        
        for section_name, pattern in section_patterns.items():
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content = match.group(1).strip()
                # Clean up the content - remove extra whitespace and newlines
                content = re.sub(r'\n+', ' ', content)
                content = re.sub(r'\s+', ' ', content)
                # Limit content length for better performance and readability
                if len(content) > 1000:
                    content = content[:1000] + "..."
                sections[section_name] = content
        
        # If we couldn't parse sections, use the whole response as summary
        if not sections:
            sections['summary'] = ai_response
        
        return sections
    
    def _estimate_project_hours(self, document_text: str, application_type: str) -> float:
        """Estimate project hours based on document complexity and type"""
        
        text_lower = document_text.lower()
        
        # Base hours by application type
        base_hours = {
            'web': 200,
            'mobile': 300,
            'both': 400,
            'desktop': 250,
            'api': 150
        }
        
        hours = base_hours.get(application_type, 200)
        
        # Complexity indicators that add hours
        complexity_indicators = {
            'database': 40,
            'user management': 30,
            'authentication': 25,
            'payment': 50,
            'integration': 35,
            'api': 30,
            'reporting': 25,
            'dashboard': 35,
            'admin': 20,
            'security': 30,
            'compliance': 40,
            'mobile': 50,
            'real-time': 40,
            'notification': 15,
            'search': 20,
            'analytics': 30,
            'blockchain': 100,
            'ai': 60,
            'machine learning': 80
        }
        
        # Add hours based on mentioned features
        for feature, additional_hours in complexity_indicators.items():
            if feature in text_lower:
                hours += additional_hours
        
        # Document length factor (longer documents = more complex requirements)
        length_factor = min(len(document_text) / 10000, 2.0)  # Cap at 2x
        hours *= (1 + length_factor)
        
        return round(hours, 1)
    
    def _calculate_project_cost(self, hours: float, rate_card: dict) -> float:
        """Calculate project cost based on hours and rate card"""
        
        # Simple calculation using average rate
        average_rate = sum(rate_card.values()) / len(rate_card)
        return round(hours * average_rate, 2)
    
    def _assess_complexity(self, document_text: str, estimated_hours) -> str:
        """Assess project complexity based on content and estimated hours"""
        
        # Ensure estimated_hours is a number for comparison
        try:
            hours = float(estimated_hours)
        except (ValueError, TypeError):
            hours = 200.0  # Default fallback
        
        text_lower = document_text.lower()
        
        # High complexity indicators
        high_complexity_keywords = [
            'enterprise', 'complex', 'integration', 'scalable', 'real-time',
            'blockchain', 'ai', 'machine learning', 'microservices', 'distributed'
        ]
        
        # Medium complexity indicators  
        medium_complexity_keywords = [
            'dashboard', 'reporting', 'user management', 'authentication',
            'payment', 'api', 'mobile', 'responsive'
        ]
        
        high_score = sum(1 for keyword in high_complexity_keywords if keyword in text_lower)
        medium_score = sum(1 for keyword in medium_complexity_keywords if keyword in text_lower)
        
        # Determine complexity based on keywords and hours
        if high_score >= 3 or hours > 400:
            return 'High'
        elif medium_score >= 3 or hours > 200:
            return 'Medium'
        else:
            return 'Low'

    def _detect_features(self, text_lower: str) -> List[str]:
        """Detect key features mentioned in the document"""
        
        feature_keywords = {
            "user authentication": ["login", "auth", "user account", "registration"],
            "database management": ["database", "data storage", "data management"],
            "reporting and analytics": ["report", "analytics", "dashboard", "metrics"],
            "API integration": ["api", "integration", "web service", "third-party"],
            "mobile access": ["mobile", "smartphone", "tablet", "app"],
            "payment processing": ["payment", "billing", "transaction", "checkout"],
            "notification system": ["notification", "email", "alert", "messaging"],
            "search functionality": ["search", "filter", "query", "find"],
            "file management": ["upload", "download", "file", "document"],
            "security features": ["security", "encryption", "secure", "protection"]
        }
        
        detected_features = []
        for feature, keywords in feature_keywords.items():
            if any(keyword in text_lower for keyword in keywords):
                detected_features.append(feature)
        
        return detected_features

    def _assess_complexity_by_features(self, word_count: int, features: List[str]) -> str:
        """Assess project complexity based on document analysis"""
        
        complexity_score = 0
        
        # Word count factor
        if word_count > 2000:
            complexity_score += 3
        elif word_count > 1000:
            complexity_score += 2
        elif word_count > 500:
            complexity_score += 1
        
        # Feature count factor
        feature_count = len(features)
        if feature_count > 8:
            complexity_score += 3
        elif feature_count > 5:
            complexity_score += 2
        elif feature_count > 3:
            complexity_score += 1
        
        # Determine complexity level
        if complexity_score >= 5:
            return "high"
        elif complexity_score >= 3:
            return "medium"
        else:
            return "standard"

    async def generate_task_breakdown(self, db: Session, rfp_id: int, force_regenerate: bool = False) -> str:
        """Generate detailed task breakdown for an RFP"""
        
        print(f"🔧 Generating task breakdown for RFP {rfp_id}...")
        logger.info(f"Starting task breakdown generation for RFP {rfp_id}")
        
        try:
            # Get RFP and existing analysis
            rfp = db.query(RFP).filter(RFP.id == rfp_id).first()
            if not rfp:
                raise Exception("RFP not found")
            
            analysis = db.query(RFPAnalysis).filter(RFPAnalysis.rfp_id == rfp_id).first()
            if not analysis:
                raise Exception("No analysis found for this RFP")
            
            # Check if tasks already exist for this analysis
            existing_tasks = db.query(AnalysisTask).filter(
                AnalysisTask.analysis_id == analysis.id
            ).first()  # Just check if any task exists
            
            if existing_tasks and not force_regenerate:
                print(f"✅ Task breakdown already exists for RFP {rfp_id}")
                logger.info(f"Task breakdown already exists for RFP {rfp_id}, returning existing breakdown")
                
                # Get all existing tasks and format them as a breakdown
                all_tasks = db.query(AnalysisTask).filter(
                    AnalysisTask.analysis_id == analysis.id
                ).order_by(AnalysisTask.order_index).all()
                
                return self._format_existing_tasks_as_breakdown(db, all_tasks)
            
            if force_regenerate:
                print(f"🔄 Force regenerating task breakdown, clearing existing tasks...")
                # Delete existing tasks and subtasks (cascade will handle subtasks)
                db.query(AnalysisTask).filter(AnalysisTask.analysis_id == analysis.id).delete()
                db.commit()
            
            print(f"📝 Generating new task breakdown...")
            
            # Extract document text if available
            document_text = ""
            if rfp.document_path:
                try:
                    document_text = self.extract_text_from_file(
                        rfp.document_path, 
                        rfp.document_path.split('.')[-1].lower()
                    )
                except Exception as e:
                    print(f"⚠️ Could not extract document text: {e}")
                    document_text = "Document text not available"
            
            # Extract technology stack from analysis
            technology_stack = self._extract_technology_stack_from_analysis(analysis)
            print(f"🛠️ Technology Stack: {technology_stack}")
            
            # Generate task breakdown using AI with technology-specific context
            task_breakdown = await self._generate_detailed_task_breakdown(
                document_text, 
                analysis.summary or "", 
                technology_stack,
                analysis.complexity_level or "Medium"
            )
            
            # Parse and store tasks in database
            await self._parse_and_store_tasks(db, analysis.id, task_breakdown)
            
            print(f"✅ Task breakdown generated and stored successfully ({len(task_breakdown)} characters)")
            return task_breakdown
            
        except Exception as e:
            print(f"❌ Task breakdown generation failed: {e}")
            logger.error(f"Task breakdown generation failed for RFP {rfp_id}: {e}")
            # Return fallback task breakdown
            return self._create_fallback_task_breakdown()

    def _extract_technology_stack_from_analysis(self, analysis: RFPAnalysis) -> dict:
        """Extract technology stack information from the analysis"""
        
        print(f"🔍 Extracting technology stack from analysis...")
        
        # Parse technology stack from the analysis text
        tech_stack = {
            'frontend': 'react',
            'backend': 'python',
            'database': 'postgresql',
            'cloud': 'aws',
            'application_type': 'web'
        }
        
        if analysis.technology_stack:
            tech_text = analysis.technology_stack.lower()
            print(f"📋 Technology Stack Text: '{tech_text}'")
            
            # Frontend technologies (prioritize blockchain frontends first)
            if 'web3' in tech_text or 'dapp' in tech_text:
                tech_stack['frontend'] = 'web3-react'
            elif 'ethers' in tech_text:
                tech_stack['frontend'] = 'ethers-react'
            elif 'angular' in tech_text:
                tech_stack['frontend'] = 'angular'
            elif 'vue' in tech_text:
                tech_stack['frontend'] = 'vue'
            elif 'svelte' in tech_text:
                tech_stack['frontend'] = 'svelte'
            elif 'nextjs' in tech_text or 'next.js' in tech_text:
                tech_stack['frontend'] = 'nextjs'
            elif 'flutter' in tech_text:
                tech_stack['frontend'] = 'flutter'
            elif 'react native' in tech_text or 'react-native' in tech_text:
                tech_stack['frontend'] = 'react-native'
            elif 'react' in tech_text:
                # Check if it's Web3 React by looking for blockchain keywords
                if any(keyword in tech_text for keyword in ['ethereum', 'blockchain', 'web3', 'smart contract', 'metamask']):
                    tech_stack['frontend'] = 'web3-react'
                else:
                    tech_stack['frontend'] = 'react'
            elif 'ios' in tech_text:
                tech_stack['frontend'] = 'native-ios'
            elif 'android' in tech_text:
                tech_stack['frontend'] = 'native-android'
            
            # Backend technologies (prioritize blockchain backends)
            if 'ethereum' in tech_text or 'solidity' in tech_text:
                tech_stack['backend'] = 'ethereum'
            elif 'solana' in tech_text:
                tech_stack['backend'] = 'solana'
            elif 'polygon' in tech_text:
                tech_stack['backend'] = 'polygon'
            elif 'hyperledger' in tech_text:
                tech_stack['backend'] = 'hyperledger'
            elif 'cardano' in tech_text:
                tech_stack['backend'] = 'cardano'
            elif 'polkadot' in tech_text:
                tech_stack['backend'] = 'polkadot'
            elif 'nodejs' in tech_text or 'node.js' in tech_text:
                tech_stack['backend'] = 'nodejs'
            elif 'java' in tech_text:
                tech_stack['backend'] = 'java'
            elif '.net' in tech_text or 'c#' in tech_text:
                tech_stack['backend'] = 'csharp'
            elif 'php' in tech_text:
                tech_stack['backend'] = 'php'
            elif 'ruby' in tech_text:
                tech_stack['backend'] = 'ruby'
            elif 'go' in tech_text:
                tech_stack['backend'] = 'go'
            
            # Database technologies (prioritize decentralized storage)
            if 'ipfs' in tech_text:
                tech_stack['database'] = 'ipfs'
            elif 'arweave' in tech_text:
                tech_stack['database'] = 'arweave'
            elif 'filecoin' in tech_text:
                tech_stack['database'] = 'filecoin'
            elif 'on-chain' in tech_text or ('blockchain' in tech_text and 'storage' in tech_text):
                tech_stack['database'] = 'blockchain-native'
            elif 'mongodb' in tech_text:
                tech_stack['database'] = 'mongodb'
            elif 'mysql' in tech_text:
                tech_stack['database'] = 'mysql'
            elif 'sql server' in tech_text:
                tech_stack['database'] = 'sql-server'
            elif 'oracle' in tech_text:
                tech_stack['database'] = 'oracle'
            elif 'redis' in tech_text:
                tech_stack['database'] = 'redis'
            elif 'dynamodb' in tech_text:
                tech_stack['database'] = 'dynamodb'
            
            # Cloud platforms (prioritize blockchain infrastructure)
            if 'alchemy' in tech_text:
                tech_stack['cloud'] = 'alchemy'
            elif 'infura' in tech_text:
                tech_stack['cloud'] = 'infura'
            elif 'quicknode' in tech_text:
                tech_stack['cloud'] = 'quicknode'
            elif 'chainlink' in tech_text:
                tech_stack['cloud'] = 'chainlink'
            elif 'pinata' in tech_text:
                tech_stack['cloud'] = 'pinata'
            elif 'azure' in tech_text:
                tech_stack['cloud'] = 'azure'
            elif 'gcp' in tech_text or 'google cloud' in tech_text:
                tech_stack['cloud'] = 'gcp'
            elif 'digitalocean' in tech_text:
                tech_stack['cloud'] = 'digitalocean'
            elif 'heroku' in tech_text:
                tech_stack['cloud'] = 'heroku'
            elif 'vercel' in tech_text:
                tech_stack['cloud'] = 'vercel'
            
            # Application type
            if 'mobile' in tech_text:
                if 'web' in tech_text:
                    tech_stack['application_type'] = 'both'
                else:
                    tech_stack['application_type'] = 'mobile'
            elif 'desktop' in tech_text:
                tech_stack['application_type'] = 'desktop'
            elif 'dapp' in tech_text or 'web3' in tech_text:
                tech_stack['application_type'] = 'dapp'
        
        print(f"✅ Extracted Technology Stack: {tech_stack}")
        return tech_stack

    def _generate_technology_context(self, technology_stack: dict) -> str:
        """Generate technology-specific context and requirements"""
        
        context_parts = []
        
        # Frontend technology context
        frontend_contexts = {
            'react': "React.js development requires: Component architecture, JSX, State management (Redux/Context), React hooks, React Router, Material-UI/Styled Components, Bundle optimization with Webpack/Vite, Testing with Jest/React Testing Library.",
            'web3-react': "Web3 React DApp development requires: Web3.js/Ethers.js integration, Wallet connectivity (MetaMask/WalletConnect/Rainbow), Smart contract interactions, Blockchain state management, Gas optimization, Transaction handling, Contract event listening, Multi-chain support, ENS integration, IPFS file handling.",
            'ethers-react': "Ethers.js React DApp development requires: Ethers.js library, Provider setup, Signer management, Contract factory patterns, Event filtering, Transaction confirmation handling, Error handling for blockchain operations, Gas estimation, Network switching.",
            'angular': "Angular development requires: TypeScript, Component/Service architecture, Angular CLI, RxJS observables, Angular Material, Reactive forms, Lazy loading modules, NgRx for state management, Karma/Jasmine testing.",
            'vue': "Vue.js development requires: Vue CLI/Vite, Component composition, Vuex/Pinia state management, Vue Router, Vuetify/Quasar UI framework, Single-file components, Testing with Vue Test Utils.",
            'flutter': "Flutter development requires: Dart programming, Widget architecture, State management (Provider/Bloc/Riverpod), Flutter packages, Platform-specific integrations, Testing framework.",
            'react-native': "React Native development requires: JavaScript/TypeScript, Native modules, React Navigation, State management, Platform-specific code (iOS/Android), Testing with Detox.",
            'nextjs': "Next.js development requires: Server-side rendering, Static generation, API routes, File-based routing, Image optimization, Performance optimization, Deployment on Vercel."
        }
        
        # Backend technology context
        backend_contexts = {
            'python': "Python backend development requires: FastAPI/Django/Flask framework, SQLAlchemy ORM, Pydantic models, Alembic migrations, Celery for async tasks, Testing with pytest, Virtual environments.",
            'nodejs': "Node.js backend development requires: Express.js/Fastify, TypeScript, ORM (Prisma/TypeORM/Sequelize), JWT authentication, Middleware setup, Testing with Jest/Mocha, NPM package management.",
            'java': "Java backend development requires: Spring Boot, Spring Security, JPA/Hibernate, Maven/Gradle, JUnit testing, Docker containerization, RESTful API design.",
            'ethereum': "Ethereum development requires: Solidity smart contracts, Hardhat/Truffle development environment, OpenZeppelin libraries, Gas optimization, Security audits, Web3 integration.",
            'solana': "Solana development requires: Rust programming, Anchor framework, Program development, Token/NFT standards, Solana Web3.js, Testing with Solana Test Validator.",
            'csharp': ".NET development requires: ASP.NET Core, Entity Framework Core, Identity management, Dependency injection, MediatR pattern, xUnit testing, NuGet packages."
        }
        
        # Database technology context
        database_contexts = {
            'postgresql': "PostgreSQL database requires: Schema design, Indexing strategies, Query optimization, Migrations, Connection pooling, Backup/restore procedures, ACID compliance.",
            'mongodb': "MongoDB database requires: Document schema design, Aggregation pipelines, Indexing, Sharding strategies, Replica sets, GridFS for file storage.",
            'mysql': "MySQL database requires: Relational schema design, Stored procedures, Triggers, Indexing optimization, Replication setup, InnoDB engine configuration.",
            'ipfs': "IPFS storage requires: Distributed file system setup, Content addressing, IPFS node configuration, Pinning strategies, Gateway setup.",
            'blockchain-native': "On-chain storage requires: Smart contract storage optimization, Gas cost considerations, Data compression techniques, Event logging strategies."
        }
        
        # Cloud platform context
        cloud_contexts = {
            'aws': "AWS deployment requires: EC2/ECS/Lambda services, RDS/DynamoDB databases, S3 storage, CloudFront CDN, Route 53 DNS, IAM security, CloudFormation/CDK infrastructure.",
            'azure': "Azure deployment requires: App Service/Container Instances, Azure SQL/Cosmos DB, Blob storage, Application Gateway, Azure AD authentication, ARM templates.",
            'gcp': "Google Cloud deployment requires: Compute Engine/Cloud Run, Cloud SQL/Firestore, Cloud Storage, Load Balancer, Cloud IAM, Deployment Manager.",
            'vercel': "Vercel deployment requires: Next.js optimization, Serverless functions, Edge runtime, Analytics setup, Domain configuration.",
            'alchemy': "Alchemy blockchain infrastructure requires: Ethereum/Polygon node access, Enhanced APIs (NFT/Token APIs), Webhook notifications, Mempool monitoring, Gas optimization tools, Multi-chain support, IPFS gateway integration, Real-time blockchain data feeds.",
            'infura': "Infura blockchain infrastructure requires: Ethereum/IPFS node access, API key management, Rate limiting, Network switching, IPFS pinning services, Layer 2 support.",
            'quicknode': "QuickNode blockchain infrastructure requires: Multi-chain node access, GraphQL APIs, WebSocket streams, Add-on marketplace, Analytics dashboard, Custom endpoints.",
            'pinata': "Pinata IPFS infrastructure requires: IPFS pinning services, File upload APIs, Pin management, Gateway access, Metadata handling, Dedicated gateways."
        }
        
        # Add relevant contexts
        if technology_stack['frontend'] in frontend_contexts:
            context_parts.append(f"FRONTEND ({technology_stack['frontend'].upper()}) REQUIREMENTS:\n{frontend_contexts[technology_stack['frontend']]}")
        
        if technology_stack['backend'] in backend_contexts:
            context_parts.append(f"BACKEND ({technology_stack['backend'].upper()}) REQUIREMENTS:\n{backend_contexts[technology_stack['backend']]}")
        
        if technology_stack['database'] in database_contexts:
            context_parts.append(f"DATABASE ({technology_stack['database'].upper()}) REQUIREMENTS:\n{database_contexts[technology_stack['database']]}")
        
        if technology_stack['cloud'] in cloud_contexts:
            context_parts.append(f"CLOUD PLATFORM ({technology_stack['cloud'].upper()}) REQUIREMENTS:\n{cloud_contexts[technology_stack['cloud']]}")
        
        # Add blockchain-specific requirements if detected
        blockchain_backends = ['ethereum', 'solana', 'polygon', 'hyperledger', 'cardano', 'polkadot']
        if technology_stack['backend'] in blockchain_backends:
            context_parts.append("""
BLOCKCHAIN-SPECIFIC REQUIREMENTS:
- Smart contract development and testing
- Wallet integration and Web3 connectivity
- Token/cryptocurrency handling
- Gas optimization and cost management
- Security audits and best practices
- Decentralized storage integration
- Oracle integration for external data
- Multi-signature wallet support
- Cross-chain compatibility (if applicable)
- Governance token implementation (if applicable)
""")
        
        return "\n\n".join(context_parts)

    async def _generate_detailed_task_breakdown(self, document_text: str, rfp_summary: str, technology_stack: dict, complexity_level: str) -> str:
        """Generate detailed task breakdown using AI"""
        
        print(f"🤖 Generating AI-powered task breakdown...")
        
        prompt = self._create_task_breakdown_prompt(document_text, rfp_summary, technology_stack, complexity_level)
        
        # Try different AI services
        ai_services = [
            ("Gemini", self._analyze_task_breakdown_with_gemini),
            ("Hugging Face", self._analyze_task_breakdown_with_huggingface),
            ("OpenAI", self._analyze_task_breakdown_with_openai)
        ]
        
        for service_name, service_method in ai_services:
            try:
                print(f"🔄 Trying {service_name} for task breakdown...")
                result = await service_method(prompt)
                if result and len(result) > 500:  # Minimum expected length
                    print(f"✅ Task breakdown generated with {service_name}")
                    return self._format_task_breakdown_response(result)
            except Exception as e:
                print(f"⚠️ {service_name} failed: {e}")
                continue
        
        print(f"🛠️ All AI services failed, using fallback task breakdown")
        return self._create_fallback_task_breakdown(technology_stack)

    def _create_task_breakdown_prompt(self, document_text: str, rfp_summary: str, technology_stack: dict, complexity_level: str) -> str:
        """Create comprehensive technology-specific task breakdown prompt"""
        
        # Generate technology-specific context
        tech_context = self._generate_technology_context(technology_stack)
        
        return f"""
        You are a senior project manager and technical architect specializing in {technology_stack['application_type']} application development. Generate a comprehensive, technology-specific task breakdown for this RFP.

        PROJECT TECHNOLOGY STACK:
        - Frontend: {technology_stack['frontend']}
        - Backend: {technology_stack['backend']}
        - Database: {technology_stack['database']}
        - Cloud Platform: {technology_stack['cloud']}
        - Application Type: {technology_stack['application_type']}
        - Project Complexity: {complexity_level}

        {tech_context}

        RFP DOCUMENT CONTENT:
        {document_text[:3000]}...

        RFP ANALYSIS SUMMARY:
        {rfp_summary}

        TASK BREAKDOWN REQUIREMENTS:
        Generate a detailed task breakdown that covers ALL functionalities mentioned in the RFP. DO NOT limit the number of modules or tasks. Create as many modules and tasks as needed to cover everything comprehensively.

        CRITICAL REQUIREMENTS:
        1. ALL tasks must be technology-specific using the exact stack mentioned above
        2. Include ALL functionalities and features mentioned in the RFP document
        3. Create detailed subtasks with specific technical implementation steps
        4. Each task should have realistic hour estimates based on the technology complexity
        5. Include technology-specific setup, configuration, and deployment tasks
        6. Cover all aspects: frontend, backend, database, testing, security, deployment, monitoring
        7. Add blockchain-specific tasks if blockchain technologies are detected
        8. Include integration tasks for all external systems mentioned in the RFP

        TASK FORMAT:
        **Module X: [Module Name]**
        Task X.Y: [Specific Technical Task Name]
        - Description: [Detailed technical description with technology-specific implementation details]
        - Estimated Hours: [Realistic hours based on technology complexity]
        - Priority: [High/Medium/Low based on project criticality]
        - Subtasks:
          * [Subtask 1]: [Specific technical step] - [Hours] hours - [Priority]
          * [Subtask 2]: [Specific technical step] - [Hours] hours - [Priority]
          * [Continue with all necessary subtasks]

        MANDATORY MODULES (create more as needed):
        1. Project Setup & Environment Configuration
        2. {technology_stack['frontend'].title()} Frontend Development
        3. {technology_stack['backend'].title()} Backend Development  
        4. {technology_stack['database'].title()} Database Implementation
        5. Authentication & Security Implementation
        6. API Development & Integration
        7. User Interface & User Experience
        8. Testing & Quality Assurance
        9. {technology_stack['cloud'].title()} Cloud Deployment
        10. Monitoring & Maintenance
        11. [Add more modules based on RFP requirements]

        Generate the COMPLETE task breakdown covering ALL RFP requirements:
        """

    async def _analyze_task_breakdown_with_gemini(self, prompt: str) -> str:
        """Generate task breakdown using Google Gemini"""
        
        if not GOOGLE_AI_AVAILABLE or not settings.GOOGLE_API_KEY:
            raise Exception("Google Gemini not available")
            
        print(f"🔑 Configuring Gemini API...")
        genai.configure(api_key=settings.GOOGLE_API_KEY)
        
        # Try multiple model names
        model_names = ['gemini-1.5-flash', 'gemini-1.5-pro', 'gemini-pro']
        
        for model_name in model_names:
            try:
                print(f"🔍 Trying model: {model_name}")
                model = genai.GenerativeModel(model_name)
                response = model.generate_content(prompt)
                
                if response.text:
                    print(f"✅ Received response from {model_name}")
                    return response.text
                    
            except Exception as e:
                print(f"❌ Model {model_name} failed: {str(e)}")
                continue
        
        raise Exception("No Gemini models available")

    async def _analyze_task_breakdown_with_huggingface(self, prompt: str) -> str:
        """Generate task breakdown using Hugging Face"""
        
        if not REQUESTS_AVAILABLE or not settings.HUGGINGFACE_API_KEY:
            raise Exception("Hugging Face not available")
            
        # Use a text generation model for analysis
        API_URL = "https://api-inference.huggingface.co/models/microsoft/DialoGPT-large"
        headers = {"Authorization": f"Bearer {settings.HUGGINGFACE_API_KEY}"}
        
        payload = {
            "inputs": prompt,
            "parameters": {
                "max_length": 3000,
                "temperature": 0.3,
                "do_sample": True
            }
        }
        
        response = requests.post(API_URL, headers=headers, json=payload, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            if isinstance(result, list) and len(result) > 0:
                generated_text = result[0].get("generated_text", "")
                return generated_text
        
        raise Exception(f"Hugging Face API error: {response.status_code}")

    async def _analyze_task_breakdown_with_openai(self, prompt: str) -> str:
        """Generate task breakdown using OpenAI"""
        
        if not OPENAI_AVAILABLE or not settings.OPENAI_API_KEY:
            raise Exception("OpenAI not available")
            
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)
        
        response = await client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a project management expert specializing in software development task breakdown and estimation."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=3000,
            temperature=0.3
        )
        
        analysis_text = response.choices[0].message.content
        return analysis_text

    def _format_task_breakdown_response(self, ai_response: str) -> str:
        """Format and validate the task breakdown response"""
        
        print(f"📝 Formatting task breakdown response...")
        
        # Basic validation - ensure we have at least some modules and tasks
        if not ai_response or len(ai_response.strip()) < 100:
            print(f"⚠️ AI response too short or empty")
            return self._create_fallback_task_breakdown({
                'frontend': 'react',
                'backend': 'python', 
                'database': 'postgresql',
                'cloud': 'aws',
                'application_type': 'web'
            })
        
        # Check for module pattern (flexible - any number of modules)
        module_pattern = r'\*\*Module\s+\d+:'
        module_matches = re.findall(module_pattern, ai_response, re.IGNORECASE)
        
        # Check for task pattern
        task_pattern = r'Task\s+[\d.]+'
        task_matches = re.findall(task_pattern, ai_response, re.IGNORECASE)
        
        if len(module_matches) < 3 or len(task_matches) < 5:
            print(f"⚠️ Insufficient modules ({len(module_matches)}) or tasks ({len(task_matches)}) in AI response")
            return self._create_fallback_task_breakdown({
                'frontend': 'react',
                'backend': 'python', 
                'database': 'postgresql',
                'cloud': 'aws',
                'application_type': 'web'
            })
        
        print(f"✅ AI response validated: {len(module_matches)} modules, {len(task_matches)} tasks")
        return ai_response.strip()

    def _format_existing_tasks_as_breakdown(self, db: Session, tasks: List[AnalysisTask]) -> str:
        """Format existing database tasks back into breakdown string format"""
        
        print(f"📋 Formatting {len(tasks)} existing tasks as breakdown...")
        
        if not tasks:
            return self._create_fallback_task_breakdown({
                'frontend': 'react',
                'backend': 'python',
                'database': 'postgresql', 
                'cloud': 'aws',
                'application_type': 'web'
            })
        
        breakdown_text = "**TASK BREAKDOWN:**\n\n"
        
        # Group tasks by module (assuming task names follow pattern like "Module 1:", "Module 2:", etc.)
        current_module = ""
        
        for task in tasks:
            # Check if this is a module header
            if task.title.startswith("Module"):
                if current_module != task.title:
                    current_module = task.title
                    breakdown_text += f"**{task.title}**\n"
            
            # Format the task
            breakdown_text += f"Task {task.order_index}: {task.title}\n"
            breakdown_text += f"- Description: {task.description}\n"
            breakdown_text += f"- Estimated Hours: {task.estimated_hours}\n"
            breakdown_text += f"- Priority: {task.priority}\n"
            
            # Get and format subtasks
            subtasks = db.query(AnalysisSubtask).filter(
                AnalysisSubtask.task_id == task.id
            ).order_by(AnalysisSubtask.order_index).all()
            
            if subtasks:
                breakdown_text += "- Subtasks:\n"
                for subtask in subtasks:
                    breakdown_text += f"  * {subtask.title}: {subtask.description} - {subtask.estimated_hours} hours - {subtask.priority}\n"
            
            breakdown_text += "\n"
        
        return breakdown_text.strip()

    async def _parse_and_store_tasks(self, db: Session, analysis_id: int, task_breakdown: str) -> None:
        """Parse task breakdown text and store tasks/subtasks in database"""
        
        print(f"💾 Parsing and storing tasks for analysis {analysis_id}...")
        
        try:
            # Parse modules and tasks from the breakdown text
            modules = self._parse_task_breakdown_text(task_breakdown)
            
            task_order = 1
            for module in modules:
                print(f"   📁 Processing module: {module['name']}")
                
                for task_data in module['tasks']:
                    # Create the main task
                    task = AnalysisTask(
                        analysis_id=analysis_id,
                        title=task_data['name'],
                        description=task_data['description'],
                        estimated_hours=task_data['hours'],
                        priority=task_data['priority'],
                        order_index=task_order,
                        module=module['name']
                    )
                    
                    db.add(task)
                    db.flush()  # Get the task ID
                    
                    # Create subtasks
                    subtask_order = 1
                    for subtask_data in task_data['subtasks']:
                        subtask = AnalysisSubtask(
                            task_id=task.id,
                            title=subtask_data['name'],
                            description=subtask_data['description'],
                            estimated_hours=subtask_data['hours'],
                            priority=subtask_data['priority'],
                            order_index=subtask_order
                        )
                        db.add(subtask)
                        subtask_order += 1
                    
                    task_order += 1
            
            db.commit()
            print(f"✅ Successfully stored {task_order - 1} tasks in database")
            
        except Exception as e:
            print(f"❌ Failed to parse and store tasks: {e}")
            logger.error(f"Failed to parse and store tasks: {e}")
            db.rollback()
            # Don't raise exception here, just log it

    def _parse_task_breakdown_text(self, text: str) -> List[Dict]:
        """Parse task breakdown text into structured data"""
        
        modules = []
        
        # Split by modules
        module_pattern = r'\*\*Module\s+(\d+):\s+([^*]+)\*\*'
        module_matches = list(re.finditer(module_pattern, text, re.IGNORECASE))
        
        for i, module_match in enumerate(module_matches):
            module_number = module_match.group(1)
            module_name = module_match.group(2).strip()
            
            # Get the content for this module
            start_pos = module_match.end()
            end_pos = module_matches[i + 1].start() if i + 1 < len(module_matches) else len(text)
            module_content = text[start_pos:end_pos]
            
            # Parse tasks within this module
            tasks = self._parse_tasks_from_module_content(module_content)
            
            modules.append({
                'number': int(module_number),
                'name': module_name,
                'tasks': tasks
            })
        
        return modules

    def _parse_tasks_from_module_content(self, module_content: str) -> List[Dict]:
        """Parse individual tasks from module content"""
        
        tasks = []
        
        # Pattern to match tasks
        task_pattern = r'Task\s+([\d.]+):\s+([^\n]+)\n[-\s]*Description:\s+([^\n]+)\n[-\s]*Estimated Hours:\s+(\d+)\n[-\s]*Priority:\s+([^\n]+)\n[-\s]*Subtasks:\s*((?:\s*\*[^\n]+\n?)*)'
        
        task_matches = re.finditer(task_pattern, module_content, re.IGNORECASE | re.MULTILINE)
        
        for task_match in task_matches:
            task_id = task_match.group(1)
            task_name = task_match.group(2).strip()
            description = task_match.group(3).strip()
            hours = int(task_match.group(4))
            priority = task_match.group(5).strip()
            subtasks_text = task_match.group(6)
            
            # Parse subtasks
            subtasks = self._parse_subtasks_from_text(subtasks_text)
            
            tasks.append({
                'id': task_id,
                'name': task_name,
                'description': description,
                'hours': hours,
                'priority': priority,
                'subtasks': subtasks
            })
        
        return tasks

    def _parse_subtasks_from_text(self, subtasks_text: str) -> List[Dict]:
        """Parse subtasks from text"""
        
        subtasks = []
        
        # Pattern to match subtasks
        subtask_pattern = r'\*\s+([^:]+):\s+([^-]+)\s+-\s+(\d+)\s+hours\s+-\s+([^\n]+)'
        
        subtask_matches = re.finditer(subtask_pattern, subtasks_text, re.IGNORECASE)
        
        for subtask_match in subtask_matches:
            name = subtask_match.group(1).strip()
            description = subtask_match.group(2).strip()
            hours = int(subtask_match.group(3))
            priority = subtask_match.group(4).strip()
            
            subtasks.append({
                'name': name,
                'description': description,
                'hours': hours,
                'priority': priority
            })
        
        return subtasks

    def export_tasks_to_excel(self, db: Session, rfp_id: int) -> BytesIO:
        """Export tasks to Excel format with proper table structure"""
        
        print(f"📊 Exporting tasks to Excel for RFP {rfp_id}...")
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise Exception("pandas and openpyxl are required for Excel export")
        
        # Get RFP and analysis
        rfp = db.query(RFP).filter(RFP.id == rfp_id).first()
        if not rfp:
            raise Exception("RFP not found")
        
        analysis = db.query(RFPAnalysis).filter(RFPAnalysis.rfp_id == rfp_id).first()
        if not analysis:
            raise Exception("No analysis found for this RFP")
        
        # Get all tasks and subtasks
        tasks = db.query(AnalysisTask).filter(
            AnalysisTask.analysis_id == analysis.id
        ).order_by(AnalysisTask.order_index).all()
        
        if not tasks:
            raise Exception("No tasks found for export")
        
        # Prepare data for Excel
        excel_data = []
        
        for task in tasks:
            # Get subtasks for this task
            subtasks = db.query(AnalysisSubtask).filter(
                AnalysisSubtask.task_id == task.id
            ).order_by(AnalysisSubtask.order_index).all()
            
            # Add main task row
            task_row = {
                'Module': task.module or 'General',
                'Task Type': 'Main Task',
                'Task ID': task.order_index,
                'Task Name': task.title,
                'Description': task.description or '',
                'Priority': task.priority,
                'Estimated Hours': task.estimated_hours or 0,
                'Estimated Cost': task.estimated_cost or 0,
                'Complexity': task.complexity or 'Medium',
                'Category': task.category or 'Development',
                'Parent Task': '',
                'Created Date': task.created_at.strftime('%Y-%m-%d') if task.created_at else ''
            }
            excel_data.append(task_row)
            
            # Add subtask rows
            for subtask in subtasks:
                subtask_row = {
                    'Module': task.module or 'General',
                    'Task Type': 'Subtask',
                    'Task ID': f"{task.order_index}.{subtask.order_index}",
                    'Task Name': subtask.title,
                    'Description': subtask.description or '',
                    'Priority': subtask.priority or 'Medium',
                    'Estimated Hours': subtask.estimated_hours or 0,
                    'Estimated Cost': subtask.estimated_cost or 0,
                    'Complexity': 'Standard',
                    'Category': 'Development',
                    'Parent Task': task.title,
                    'Created Date': subtask.created_at.strftime('%Y-%m-%d') if subtask.created_at else ''
                }
                excel_data.append(subtask_row)
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create Excel workbook with advanced formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Task Breakdown"
        
        # Add project information header
        ws['A1'] = f"RFP: {rfp.title}"
        ws['A2'] = f"Client: {rfp.client_name}"
        ws['A3'] = f"Total Tasks: {len(tasks)}"
        ws['A4'] = f"Total Hours: {sum(task.estimated_hours or 0 for task in tasks)}"
        ws['A5'] = f"Export Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Style header
        header_font = Font(bold=True, size=14)
        for row in range(1, 6):
            ws[f'A{row}'].font = header_font
        
        # Add data starting from row 7
        start_row = 7
        
        # Add column headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Style based on task type
                if col_idx == 2:  # Task Type column
                    if value == 'Main Task':
                        cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                        cell.font = Font(bold=True)
                    else:
                        cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                
                # Style priority column
                if col_idx == 6:  # Priority column
                    if value == 'High':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        cell.font = Font(color="CC0000")
                    elif value == 'Medium':
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                        cell.font = Font(color="FF6600")
                    elif value == 'Low':
                        cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
                        cell.font = Font(color="009900")
                
                # Center align numeric columns
                if col_idx in [3, 7, 8]:  # Task ID, Hours, Cost
                    cell.alignment = Alignment(horizontal="center")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Create summary data
        module_summary = {}
        priority_summary = {'High': 0, 'Medium': 0, 'Low': 0}
        total_hours = 0
        total_cost = 0
        
        for task in tasks:
            module = task.module or 'General'
            if module not in module_summary:
                module_summary[module] = {'tasks': 0, 'hours': 0, 'cost': 0}
            
            module_summary[module]['tasks'] += 1
            module_summary[module]['hours'] += task.estimated_hours or 0
            module_summary[module]['cost'] += task.estimated_cost or 0
            
            priority_summary[task.priority] = priority_summary.get(task.priority, 0) + 1
            total_hours += task.estimated_hours or 0
            total_cost += task.estimated_cost or 0
        
        # Add summary headers
        summary_ws['A1'] = "PROJECT SUMMARY"
        summary_ws['A1'].font = Font(bold=True, size=16)
        
        summary_ws['A3'] = "Module Breakdown"
        summary_ws['A3'].font = Font(bold=True, size=12)
        
        # Module summary table
        summary_ws['A4'] = "Module"
        summary_ws['B4'] = "Tasks"
        summary_ws['C4'] = "Hours"
        summary_ws['D4'] = "Cost"
        
        row = 5
        for module, data in module_summary.items():
            summary_ws[f'A{row}'] = module
            summary_ws[f'B{row}'] = data['tasks']
            summary_ws[f'C{row}'] = data['hours']
            summary_ws[f'D{row}'] = f"${data['cost']:.2f}"
            row += 1
        
        # Priority summary
        summary_ws[f'A{row + 1}'] = "Priority Breakdown"
        summary_ws[f'A{row + 1}'].font = Font(bold=True, size=12)
        
        summary_ws[f'A{row + 2}'] = "Priority"
        summary_ws[f'B{row + 2}'] = "Count"
        
        for priority, count in priority_summary.items():
            row += 1
            summary_ws[f'A{row + 2}'] = priority
            summary_ws[f'B{row + 2}'] = count
        
        # Total summary
        summary_ws[f'A{row + 4}'] = "TOTALS"
        summary_ws[f'A{row + 4}'].font = Font(bold=True, size=12)
        summary_ws[f'A{row + 5}'] = f"Total Tasks: {len(tasks)}"
        summary_ws[f'A{row + 6}'] = f"Total Hours: {total_hours}"
        summary_ws[f'A{row + 7}'] = f"Total Cost: ${total_cost:.2f}"
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        print(f"✅ Excel export completed with {len(excel_data)} rows")
        return excel_buffer

    def _create_fallback_task_breakdown(self, technology_stack: dict) -> str:
        """Create technology-specific fallback task breakdown when AI services fail"""
        
        print(f"🛠️ Creating technology-specific fallback task breakdown...")
        print(f"📋 Tech Stack: {technology_stack}")
        
        frontend = technology_stack.get('frontend', 'react').title()
        backend = technology_stack.get('backend', 'python').title()
        database = technology_stack.get('database', 'postgresql').title()
        cloud = technology_stack.get('cloud', 'aws').upper()
        app_type = technology_stack.get('application_type', 'web')
        
        # Technology-specific configurations
        frontend_configs = {
            'react': ('React.js', 'npm install, Create React App, React Router, Redux/Context API, Material-UI/Styled Components'),
            'angular': ('Angular', 'Angular CLI, TypeScript, Angular Material, RxJS, NgRx'),
            'vue': ('Vue.js', 'Vue CLI, Vuex/Pinia, Vue Router, Vuetify/Quasar'),
            'flutter': ('Flutter', 'Flutter SDK, Dart, Provider/Bloc, Material Design'),
            'react-native': ('React Native', 'React Native CLI, Metro bundler, React Navigation'),
            'nextjs': ('Next.js', 'Next.js framework, Static generation, Server-side rendering')
        }
        
        backend_configs = {
            'python': ('Python/FastAPI', 'FastAPI framework, Pydantic, SQLAlchemy, Alembic, pytest'),
            'nodejs': ('Node.js/Express', 'Express.js, TypeScript, Prisma/TypeORM, Jest'),
            'java': ('Java/Spring', 'Spring Boot, Spring Security, JPA/Hibernate, Maven'),
            'ethereum': ('Ethereum/Solidity', 'Hardhat, OpenZeppelin, Web3.js, Smart contracts'),
            'solana': ('Solana/Rust', 'Anchor framework, Solana CLI, Token programs'),
            'csharp': ('.NET Core', 'ASP.NET Core, Entity Framework, Identity')
        }
        
        database_configs = {
            'postgresql': ('PostgreSQL', 'Relational database, ACID compliance, Advanced querying'),
            'mongodb': ('MongoDB', 'Document database, Aggregation pipelines, GridFS'),
            'mysql': ('MySQL', 'Relational database, Stored procedures, Triggers'),
            'ipfs': ('IPFS', 'Distributed storage, Content addressing, Pinning'),
            'blockchain-native': ('Blockchain Storage', 'On-chain storage, Gas optimization')
        }
        
        frontend_name, frontend_tools = frontend_configs.get(technology_stack.get('frontend', 'react'), frontend_configs['react'])
        backend_name, backend_tools = backend_configs.get(technology_stack.get('backend', 'python'), backend_configs['python'])
        database_name, database_tools = database_configs.get(technology_stack.get('database', 'postgresql'), database_configs['postgresql'])
        
        return f"""**TASK BREAKDOWN:**

**Module 1: Project Setup & {backend_name} Environment Configuration**
Task 1.1: {backend_name} Development Environment Setup
- Description: Set up {backend_name} development environment, package management, and project structure with technology-specific configurations
- Estimated Hours: 16
- Priority: High
- Subtasks:
  * Initialize {backend_name} project with {backend_tools.split(',')[0]} - 4 hours - High
  * Configure package management and dependency resolution - 4 hours - High
  * Set up development tools and IDE configuration - 4 hours - Medium
  * Configure environment variables and configuration management - 4 hours - Medium

Task 1.2: {frontend_name} Frontend Environment Setup  
- Description: Configure {frontend_name} development environment with build tools, linting, and testing framework
- Estimated Hours: 14
- Priority: High
- Subtasks:
  * Initialize {frontend_name} project with {frontend_tools.split(',')[0]} - 4 hours - High
  * Configure build system and bundling tools - 4 hours - High
  * Set up linting, formatting, and code quality tools - 3 hours - Medium
  * Configure testing framework and development scripts - 3 hours - Medium

Task 1.3: {database_name} Database Environment Setup
- Description: Configure {database_name} database environment, connection pooling, and migration system
- Estimated Hours: 12
- Priority: High
- Subtasks:
  * Install and configure {database_name} database server - 4 hours - High
  * Set up database connection pooling and optimization - 4 hours - Medium
  * Configure migration system and version control - 4 hours - Medium

Task 1.4: {cloud} Cloud Infrastructure Planning
- Description: Design {cloud} cloud infrastructure, resource planning, and deployment strategy
- Estimated Hours: 18
- Priority: Medium
- Subtasks:
  * {cloud} account setup and resource planning - 6 hours - Medium
  * Infrastructure as Code (IaC) configuration - 8 hours - Medium
  * CI/CD pipeline design and implementation - 4 hours - Medium

**Module 2: Authentication & Security Implementation**
Task 2.1: User Authentication System
- Description: Implement secure user registration, login, password management, and session handling mechanisms
- Estimated Hours: 24
- Priority: High
- Subtasks:
  * User registration and email verification system - 8 hours - High
  * Secure login and password management - 8 hours - High
  * Session management and token-based authentication - 8 hours - High

Task 2.2: Authorization & Access Control
- Description: Implement role-based access control, permissions system, and secure API endpoints
- Estimated Hours: 18
- Priority: High
- Subtasks:
  * Role-based access control implementation - 8 hours - High
  * Permission system and middleware development - 6 hours - High
  * API security and endpoint protection - 4 hours - Medium

Task 2.3: Security Hardening
- Description: Implement security best practices, encryption, input validation, and vulnerability protection
- Estimated Hours: 16
- Priority: Medium
- Subtasks:
  * Data encryption and secure communication protocols - 6 hours - High
  * Input validation and sanitization systems - 5 hours - Medium
  * Security testing and vulnerability assessment - 5 hours - Medium

**Module 3: Database Design & Implementation**
Task 3.1: Database Schema Development
- Description: Design and implement comprehensive database schema with relationships, indexes, and optimization
- Estimated Hours: 20
- Priority: High
- Subtasks:
  * Entity relationship design and normalization - 8 hours - High
  * Database schema implementation and migration scripts - 8 hours - High
  * Index optimization and performance tuning - 4 hours - Medium

Task 3.2: Data Access Layer
- Description: Implement data access patterns, ORM configuration, and database abstraction layers
- Estimated Hours: 16
- Priority: High
- Subtasks:
  * ORM setup and model configuration - 6 hours - High
  * Repository pattern and data access implementation - 6 hours - High
  * Database connection pooling and optimization - 4 hours - Medium

Task 3.3: Data Migration & Backup
- Description: Implement data migration tools, backup strategies, and database maintenance procedures
- Estimated Hours: 12
- Priority: Medium
- Subtasks:
  * Data migration scripts and version control - 5 hours - Medium
  * Automated backup and recovery procedures - 4 hours - Medium
  * Database monitoring and maintenance tools - 3 hours - Low

**Module 4: Frontend User Interface Development**
Task 4.1: Core UI Components
- Description: Develop reusable UI components, design system, and responsive layout framework
- Estimated Hours: 28
- Priority: High
- Subtasks:
  * Component library and design system development - 10 hours - High
  * Responsive layout and grid system implementation - 8 hours - High
  * Navigation and routing system development - 10 hours - High

Task 4.2: User Interface Implementation
- Description: Implement main application screens, forms, and interactive elements with user experience optimization
- Estimated Hours: 32
- Priority: High
- Subtasks:
  * Main dashboard and navigation interfaces - 12 hours - High
  * Forms and data input interfaces - 10 hours - High
  * Interactive features and user feedback systems - 10 hours - Medium

Task 4.3: Frontend State Management
- Description: Implement state management, data synchronization, and client-side caching mechanisms
- Estimated Hours: 20
- Priority: Medium
- Subtasks:
  * Global state management implementation - 8 hours - High
  * Data synchronization and real-time updates - 8 hours - Medium
  * Client-side caching and optimization - 4 hours - Medium

**Module 5: Backend API Development**
Task 5.1: Core API Framework
- Description: Develop RESTful API structure, middleware, error handling, and documentation framework
- Estimated Hours: 24
- Priority: High
- Subtasks:
  * API framework setup and routing configuration - 8 hours - High
  * Middleware development and error handling - 8 hours - High
  * API documentation and testing framework - 8 hours - Medium

Task 5.2: Business Logic Implementation
- Description: Implement core business logic, data validation, and service layer architecture
- Estimated Hours: 36
- Priority: High
- Subtasks:
  * Core business logic and service implementation - 16 hours - High
  * Data validation and business rule enforcement - 10 hours - High
  * Service layer architecture and dependency injection - 10 hours - Medium

Task 5.3: API Integration & Third-party Services
- Description: Integrate external APIs, payment processing, and third-party service connections
- Estimated Hours: 20
- Priority: Medium
- Subtasks:
  * External API integration and data mapping - 8 hours - Medium
  * Payment processing and financial transaction handling - 8 hours - High
  * Third-party service authentication and error handling - 4 hours - Medium

**Module 6: Integration & Data Management**
Task 6.1: System Integration
- Description: Implement frontend-backend integration, real-time communication, and data synchronization
- Estimated Hours: 18
- Priority: High
- Subtasks:
  * Frontend-backend API integration - 8 hours - High
  * Real-time communication and WebSocket implementation - 6 hours - Medium
  * Data synchronization and conflict resolution - 4 hours - Medium

Task 6.2: Data Processing & Analytics
- Description: Implement data processing pipelines, reporting systems, and analytics functionality
- Estimated Hours: 24
- Priority: Medium
- Subtasks:
  * Data processing and transformation pipelines - 10 hours - Medium
  * Reporting system and dashboard analytics - 10 hours - Medium
  * Performance monitoring and data insights - 4 hours - Low

Task 6.3: File Management & Storage
- Description: Implement file upload, storage management, and content delivery systems
- Estimated Hours: 16
- Priority: Medium
- Subtasks:
  * File upload and validation systems - 6 hours - Medium
  * Cloud storage integration and management - 6 hours - Medium
  * Content delivery and optimization - 4 hours - Low

**Module 7: Testing & Quality Assurance**
Task 7.1: Automated Testing Implementation
- Description: Develop comprehensive testing suite including unit tests, integration tests, and end-to-end testing
- Estimated Hours: 28
- Priority: High
- Subtasks:
  * Unit testing framework and test case development - 12 hours - High
  * Integration testing and API test automation - 10 hours - High
  * End-to-end testing and user journey validation - 6 hours - Medium

Task 7.2: Code Quality & Performance Testing
- Description: Implement code quality checks, performance testing, and optimization procedures
- Estimated Hours: 20
- Priority: Medium
- Subtasks:
  * Code quality analysis and linting automation - 6 hours - Medium
  * Performance testing and load testing implementation - 10 hours - Medium
  * Security testing and vulnerability scanning - 4 hours - High

Task 7.3: Bug Tracking & Quality Assurance
- Description: Establish bug tracking, quality assurance processes, and user acceptance testing procedures
- Estimated Hours: 16
- Priority: Medium
- Subtasks:
  * Bug tracking and issue management system setup - 5 hours - Medium
  * Quality assurance process and testing protocols - 6 hours - Medium
  * User acceptance testing and feedback integration - 5 hours - Low

**Module 8: Deployment & Production Setup**
Task 8.1: Production Environment Setup
- Description: Configure production infrastructure, deployment pipelines, and monitoring systems
- Estimated Hours: 24
- Priority: High
- Subtasks:
  * Production server configuration and optimization - 10 hours - High
  * Deployment automation and CI/CD pipeline setup - 8 hours - High
  * Monitoring and logging system implementation - 6 hours - Medium

Task 8.2: Go-Live & Launch Preparation
- Description: Execute production deployment, perform final testing, and prepare launch procedures
- Estimated Hours: 16
- Priority: High
- Subtasks:
  * Production deployment and system validation - 8 hours - High
  * Final testing and performance verification - 5 hours - High
  * Launch preparation and rollback procedures - 3 hours - Medium

Task 8.3: Post-Launch Support & Maintenance
- Description: Establish maintenance procedures, support systems, and continuous improvement processes
- Estimated Hours: 12
- Priority: Low
- Subtasks:
  * Support documentation and maintenance procedures - 4 hours - Medium
  * Performance monitoring and optimization setup - 4 hours - Medium
  * Continuous improvement and feedback systems - 4 hours - Low"""

# Create singleton instance
document_analysis_service = DocumentAnalysisService() 